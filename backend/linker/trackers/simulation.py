import csv
import datetime
import gzip
import re
import zoneinfo
import json
from pathlib import Path
from typing import Optional

from dateutil.parser import isoparse
from django.conf import settings
from django.contrib.gis.gdal import DataSource, GDALException
from django.contrib.gis.geos import GEOSGeometry
from django.utils.timezone import now
from openpyxl.reader.excel import load_workbook

from .constants import SETTING_SIMULATION_START
from .geodynamics import import_geodynamics_data_batch
from .models import TrackerLog, Tracker
from ..config.models import Setting
from ..map.models import Tocht, Weide, Fiche, Zijweg
from ..people.constants import MemberType, Direction
from ..people.models import Team, OrganizationMember, ContactPerson
from ..tracing.models import CheckpointLog

SIMULATION_START = datetime.datetime(year=2023, month=4, day=29, hour=12, tzinfo=zoneinfo.ZoneInfo('Europe/Brussels'))


def restart_simulation():
    timestamp = now()
    TrackerLog.objects.filter(fetch_datetime__gt=SIMULATION_START).delete()
    CheckpointLog.objects.filter(timestamp__gt=SIMULATION_START).delete()
    for tracker in Tracker.objects.all():
        tracker.last_log = tracker.tracker_logs.order_by('-gps_datetime').first()
        tracker.save()
    Setting.objects.update_or_create(key=SETTING_SIMULATION_START, defaults={'value': timestamp.isoformat()})


def _get_timestamp_from_file_name(file: Path) -> datetime.datetime:
    return datetime.datetime(1970, 1, 1, tzinfo=zoneinfo.ZoneInfo('UTC')) + datetime.timedelta(
        seconds=(int(file.name.replace('.json.gz', '')))
    )


def simulate_download_tracker_data(until: Optional[datetime.datetime] = None):
    if until is None:
        try:
            start_timestamp_str = Setting.get_value_for_key(SETTING_SIMULATION_START)
            start_timestamp = isoparse(start_timestamp_str)
            until = SIMULATION_START + (now() - start_timestamp)
        except KeyError:
            until = SIMULATION_START

    files_to_do = list((Path(settings.SIMULATION_PATH) / 'geodynamics').glob('*.json.gz'))
    try:
        latest_datetime = TrackerLog.objects.latest('fetch_datetime').fetch_datetime
        files_to_do = [
            filename for filename in files_to_do if latest_datetime < _get_timestamp_from_file_name(filename)
        ]
    except TrackerLog.DoesNotExist:
        pass

    files_to_do = [filename for filename in files_to_do if _get_timestamp_from_file_name(filename) <= until]
    files_to_do.sort(key=lambda filename: filename.name)

    batch_size = 100

    for index in range(0, len(files_to_do), batch_size):
        all_data = []
        print(_get_timestamp_from_file_name(files_to_do[index]))
        for filename in files_to_do[index : index + batch_size]:
            with gzip.open(filename, 'rb') as file:
                all_data.append((_get_timestamp_from_file_name(filename), json.load(file)))

        import_geodynamics_data_batch(all_data)

    for tracker in Tracker.objects.all():
        tracker.last_log = tracker.tracker_logs.order_by('-gps_datetime').first()
        tracker.save()


def couple_trackers():
    for tracker in Tracker.objects.all():
        if hasattr(tracker, 'team'):
            continue
        code = tracker.last_log.code
        if code.startswith('RK'):
            continue
        if not code[0] == 'R' and not code[0] == 'B':
            continue
        team_id = int(code[1:])
        team = Team.objects.get(number=team_id)
        team.tracker = tracker
        team.save()

    for tracker in Tracker.objects.all():
        if hasattr(tracker, 'team'):
            continue
        code = tracker.last_log.code
        if len(code) == 0 or not code.isnumeric():
            continue
        team_id = int(code)
        team = Team.objects.filter(number=team_id).first()
        if team and team.tracker is None:
            team.tracker = tracker
            team.save()

    OrganizationMember.objects.all().delete()
    for tracker in Tracker.objects.all():
        code = tracker.last_log.code
        member = OrganizationMember.objects.filter(code=code).first()
        if member is not None:
            member.tracker = tracker
            member.save()


def import_gpkg(filename: Path):
    ds = DataSource(filename)

    for feature in ds['Tocht']:
        identifier = str(feature['name'])[0].upper()
        Tocht.objects.update_or_create(
            identifier=identifier,
            order=ord(identifier) - ord('A'),
            defaults=dict(route=GEOSGeometry(feature.geom[0].ewkt)),
        )

    for feature in ds['Weides']:
        name = str(feature['name'])

        if name.lower() == 'basis':
            continue

        tocht = Tocht.objects.get(identifier=name[0].upper())
        Weide.objects.update_or_create(tocht=tocht, defaults=dict(polygon=GEOSGeometry(feature.geom[0].ewkt)))

    for feature in ds['Fiches']:
        name = str(feature['name'])
        tocht = Tocht.objects.get(identifier=name[0].upper())
        order = int(name[1:])
        Fiche.objects.update_or_create(
            tocht=tocht,
            order=order,
            defaults=dict(point=GEOSGeometry(feature.geom.ewkt)),
        )

    Zijweg.objects.all().delete()
    for feature in ds['Zijwegen']:
        try:
            Zijweg.objects.create(geom=feature.geom[0].ewkt)
        except GDALException:
            pass


def import_groepen_en_deelnemers(filename: Path):
    wb = load_workbook(str(filename), read_only=True, data_only=True)

    teams = wb['LIJST Inschrijvingen groepen']

    first_row = [cell.value for cell in teams['1']]
    id_col = first_row.index('G_ID')
    name_col = first_row.index('Teamnaam')
    chiro_col = first_row.index('Chirogroep')
    weide_col = first_row.index('StartWei')
    richting_col = first_row.index('Richting')

    for row in teams.iter_rows(min_row=2):
        id = int(row[id_col].value)
        if id == 0:
            continue
        name = row[name_col].value
        if name.lower() == 'annulatie' or name == '#N/A':
            continue
        name = name[0].upper() + name[1:]
        chiro = row[chiro_col].value
        weide = Weide.objects.get(tocht__identifier=row[weide_col].value[0].upper())
        richting = Direction(row[richting_col].value)

        Team.objects.update_or_create(
            number=id,
            defaults=dict(direction=richting, name=name, chiro=chiro, start_weide_1=weide),
        )

    personen = wb['LIJST Inschrijvingen personen']

    first_row = [cell.value for cell in personen['1']]
    id_col = first_row.index('G_ID')
    volgnr_col = first_row.index('Volgnr')
    name_col = first_row.index('Naam')
    email_col = first_row.index('e-mail')
    phone_col = first_row.index('GSM')
    team_col = first_row.index('Teamnaam')

    for row in personen.iter_rows(min_row=2):
        id = int(row[id_col].value)
        if id == 0:
            continue
        if row[team_col].value == '#N/A' or row[team_col].value.lower() == 'annulatie':
            continue
        name = str(row[name_col].value.title())
        email = str(row[email_col].value)
        phone = ''.join(c for c in str(row[phone_col].value) if c.isdigit() or c == '+')
        is_leader = int(row[volgnr_col].value) == 1

        if len(email) <= 1:
            email = None
        if len(phone) <= 1:
            phone = None

        if phone is not None and phone.startswith('4'):
            phone = '+32' + phone
        if phone is not None and phone.startswith('032'):
            phone = '+' + phone[1:]

        if phone is not None and not re.match(r'(\+?32|0032|0)4\d{8}', phone):
            print(f'Warning: phone number not correct: {phone} from {name}')

        team = Team.objects.get(number=id)
        ContactPerson.objects.create(
            name=name,
            phone_number=phone,
            email_address=email,
            team=team,
            is_leader=is_leader,
        )


def import_organization_members(filename: Path):
    with filename.open('r') as f:
        data = list(csv.DictReader(f))
    OrganizationMember.objects.all().delete()
    for line in data:
        OrganizationMember.objects.create(
            tracker=None,
            name=line['name'],
            code=line['code'],
            member_type=MemberType(line['member_type']),
        )
