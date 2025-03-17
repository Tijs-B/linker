import csv
import json
import re
from pathlib import Path
from typing import Any

from django.contrib.gis.gdal import DataSource, GDALException
from django.contrib.gis.geos import GEOSGeometry
from openpyxl import load_workbook
from requests import get

from linker.map.models import Tocht, Basis, Weide, Fiche, Zijweg
from linker.people.constants import Direction, MemberType
from linker.people.models import Team, ContactPerson, OrganizationMember


def import_gpkg(filename: Path):
    ds = DataSource(filename)

    for feature in ds['Tocht']:
        identifier = str(feature['name'])[0].upper()
        Tocht.objects.update_or_create(
            identifier=identifier,
            order=ord(identifier) - ord('A') + 1,
            defaults=dict(route=GEOSGeometry(feature.geom[0].ewkt)),
        )

    for feature in ds['Weides']:
        name = str(feature['name'])
        geometry = GEOSGeometry(feature.geom[0].ewkt)

        if name.lower() == 'basis':
            basis = Basis.objects.first()
            if basis is None:
                Basis.objects.create(point=geometry.centroid)
            else:
                basis.point = geometry.centroid
                basis.save()

        tocht = Tocht.objects.get(identifier=name[0].upper())
        Weide.objects.update_or_create(tocht=tocht, defaults=dict(polygon=geometry))

    for feature in ds['Fiches']:
        name = str(feature['name'])
        tocht = Tocht.objects.get(identifier=name[0].upper())
        order = int(name[1:])
        Fiche.objects.update_or_create(
            tocht=tocht,
            order=order,
            defaults=dict(point=GEOSGeometry(feature.geom.ewkt)),
        )

    Zijweg.objects.all().delete()
    for feature in ds['Zijwegen']:
        try:
            Zijweg.objects.create(geom=feature.geom[0].ewkt)
        except GDALException:
            pass


def _import_geojson_tochten(geojson: dict[str, Any]) -> None:
    for feature in geojson['features']:
        if feature.get('geometry') is not None:
            route = GEOSGeometry(json.dumps(feature['geometry']))
            letter = feature['properties']['letter']
            Tocht.objects.update_or_create(
                identifier=letter,
                order=ord(letter) - ord('A') + 1,
                defaults=dict(route=route),
            )


def _import_geojson_weides(geojson: dict[str, Any]) -> None:
    for feature in geojson['features']:
        if feature.get('geometry') is not None:
            polygon = GEOSGeometry(json.dumps(feature['geometry']))
            letter = feature['properties']['letter'][0].upper()
            name = feature['properties']['name']
            if letter == 'X':
                basis = Basis.objects.first()
                if basis is None:
                    Basis.objects.create(point=polygon.centroid)
                else:
                    basis.point = polygon.centroid
                    basis.save()
                continue

            tocht = Tocht.objects.get(identifier=letter)
            Weide.objects.update_or_create(tocht=tocht, defaults=dict(polygon=polygon, identifier=letter, name=name))


def _import_geojson_fiches(geojson: dict[str, Any]) -> None:
    for feature in geojson['features']:
        if feature.get('geometry') is not None:
            point = GEOSGeometry(json.dumps(feature['geometry']))
            name = feature['properties']['name']
            tocht = Tocht.objects.get(identifier=name[0].upper())
            order = int(name[1:])
            Fiche.objects.update_or_create(
                tocht=tocht,
                order=order,
                defaults=dict(point=point),
            )


def _import_geojson_zijwegen(geojson: dict[str, Any]) -> None:
    Zijweg.objects.all().delete()
    for feature in geojson['features']:
        if feature.get('geometry') is not None:
            line = GEOSGeometry(json.dumps(feature['geometry']))
            Zijweg.objects.create(geom=line)


def import_geoserver(base_url: str):
    params = {
        'service': 'WFS',
        'version': '1.0.0',
        'request': 'GetFeature',
        'typeName': 'Chirolink:tochten_2024',
        'maxFeatures': 1000,
        'outputFormat': 'application/json',
    }
    tochten = get(base_url + '/geoserver/Chirolink/ows', params={**params, 'typeName': 'Chirolink:tochten_2024'})
    _import_geojson_tochten(tochten.json())

    weides = get(base_url + '/geoserver/Chirolink/ows', params={**params, 'typeName': 'Chirolink:weides_2024'})
    _import_geojson_weides(weides.json())

    fiches = get(base_url + '/geoserver/Chirolink/ows', params={**params, 'typeName': 'Chirolink:fiches_2024'})
    _import_geojson_fiches(fiches.json())

    zijwegen = get(base_url + '/geoserver/Chirolink/ows', params={**params, 'typeName': 'Chirolink:zijwegen_2024'})
    _import_geojson_zijwegen(zijwegen.json())


def import_featureserv(base_url: str):
    params = {'limit': 10_000}
    stripped = base_url.rstrip('/')

    tochten = get(stripped + '/collections/public.tochten_2025/items.json', params=params)
    _import_geojson_tochten(tochten.json())

    weides = get(stripped + '/collections/public.weides_2025/items.json', params=params)
    _import_geojson_weides(weides.json())

    fiches = get(stripped + '/collections/public.fiches_2025/items.json', params=params)
    _import_geojson_fiches(fiches.json())

    zijwegen = get(stripped + '/collections/public.zijwegen_2025/items.json', params=params)
    _import_geojson_zijwegen(zijwegen.json())


def import_groepen_en_deelnemers(filename: Path):
    wb = load_workbook(str(filename), read_only=True, data_only=True)

    teams = wb['LIJST Inschrijvingen groepen']

    first_row = [cell.value for cell in teams['1']]
    g_id_Col = first_row.index('G_ID')
    name_col = first_row.index('Teamnaam')
    chiro_col = first_row.index('Chirogroep')
    richting_col = first_row.index('Richting')

    for row in teams.iter_rows(min_row=2):
        g_id = int(row[g_id_Col].value)
        if g_id == 0:
            continue
        name = row[name_col].value
        if name.lower() == 'annulatie' or name == '#N/A':
            continue
        name = name[0].upper() + name[1:]
        chiro = row[chiro_col].value
        richting = Direction(row[richting_col].value)

        Team.objects.update_or_create(
            number=g_id,
            defaults=dict(direction=richting, name=name, chiro=chiro),
        )

    personen = wb['LIJST Inschrijvingen personen']

    first_row = [cell.value for cell in personen['1']]
    g_id_Col = first_row.index('G_ID')
    volgnr_col = first_row.index('Volgnr')
    name_col = first_row.index('Naam')
    email_col = first_row.index('e-mail')
    phone_col = first_row.index('GSM')
    team_col = first_row.index('Teamnaam')

    for row in personen.iter_rows(min_row=2):
        g_id = int(row[g_id_Col].value)
        if g_id == 0:
            continue
        if row[team_col].value == '#N/A' or row[team_col].value.lower() == 'annulatie':
            continue
        if row[name_col].value is None:
            print(f'Warning: empty name in row {row}. Skipping')
            continue
        name = str(row[name_col].value.title())
        email = str(row[email_col].value)
        phone = ''.join(c for c in str(row[phone_col].value) if c.isdigit() or c == '+')
        is_favorite = int(row[volgnr_col].value) == 1

        if len(email) <= 1:
            email = None
        if len(phone) <= 1:
            phone = None

        if phone is not None:
            phone = phone.replace(' ', '')
            phone = re.sub(r'^(?:\+?32|0032|0)?4', '+324', phone)

        if phone is not None and not re.match(r'\+324\d{8}', phone):
            print(f'Warning: phone number not correct: {phone} from {name}. Skipping adding.')
            phone = None

        team = Team.objects.get(number=g_id)
        ContactPerson.objects.create(
            name=name, phone_number=phone, email_address=email, team=team, is_favorite=is_favorite
        )


def import_organization_members(filename: Path):
    with filename.open('r') as f:
        data = list(csv.DictReader(f))
    OrganizationMember.objects.all().delete()
    for line in data:
        OrganizationMember.objects.create(
            tracker=None,
            name=line['name'],
            code=line['code'],
            member_type=MemberType(line['member_type']),
        )
